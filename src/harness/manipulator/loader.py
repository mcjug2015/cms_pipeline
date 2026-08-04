import argparse
import datetime
import os
import sys
import tempfile
import uuid
from abc import ABC, abstractmethod
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from pyspark.sql import SparkSession

from src import custom_logging
from src.crutch_migrations.run_crutch_migrations import (
    get_ascending_letters_within_minute,
)
from src.harness.manipulator.unwrapper import Unwrapper
from src.spark_utils import get_spark
from src.utils import convert_to_key, download_s3_zip

logger = custom_logging.setup_logging().getLogger(__name__)


class AbstractLoader(ABC, Unwrapper):

    @abstractmethod
    def get_sheet_name(self):
        pass

    @abstractmethod
    def get_first_header_cell_val(self):
        pass

    @abstractmethod
    def get_s3_zip_uri(self):
        pass

    def parse_sheet(
        self,
        xlsx_path: str,
    ) -> Tuple[int, List[Dict[str, Any]]]:
        workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
        worksheet = workbook[self.get_sheet_name()]

        data_rows: List[Dict[str, Any]] = []
        col_index_to_header_col_name = {}
        for header_row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            cells = [c for c in row if c is not None and str(c).strip() != ""]
            if cells[0] == self.get_first_header_cell_val():
                for idx, header_cell in enumerate(cells):
                    col_index_to_header_col_name[idx] = header_cell
                break

        for idx, row in enumerate(
            worksheet.iter_rows(min_row=header_row_idx + 2, values_only=True)
        ):
            cells = [c for c in row if c is not None and str(c).strip() != ""]
            if (
                len(cells) > 0
                and len(str(cells[0]).strip()) > 0
                and str(cells[0]).strip() != "BLANK"
            ):
                record = {}
                for idx, value_cell in enumerate(cells):
                    record[str(col_index_to_header_col_name[idx])] = str(value_cell)
                data_rows.append(record)

        return workbook.index(worksheet), data_rows

    def insert_kvp_rows(
        self,
        spark: SparkSession,
        cat: str,
        schema: str,
        load_id: str,
        zip_name: str,
        unzipped_name: str,
        sheet_name: str,
        sheet_index: int,
        data_rows: List[Dict[str, str]],
    ) -> None:
        if not data_rows:
            return
        val_strs = []
        for data_row in data_rows:
            for the_key, the_val in data_row.items():
                table_key_simple = convert_to_key(the_key)
                # literal strs since args don't work with local spark used for testing.
                val_strs.append(
                    f"('{load_id}', '{zip_name}', '{unzipped_name}', '{sheet_name}', {sheet_index}"
                    f", '{the_key}', '{table_key_simple}'"
                    f", '{the_val}', current_timestamp(), current_timestamp())"
                )
        spark.sql(
            f"""
            insert into {cat}.{schema}.open_cms_data_kvp
            (load_id, zip_name, unzipped_name, sheet_name, sheet_index, table_key, table_key_simple
             , table_val, created_at, updated_at)
            values {", ".join(val_strs)}
            """
        )

    def load(self, spark: SparkSession, cat: str, schema: str) -> Dict[str, int]:
        with tempfile.TemporaryDirectory(prefix="cms_dl_") as tmp_dir:
            zip_path = download_s3_zip(spark, self.get_s3_zip_uri(), tmp_dir)
            with self.unwrap(zip_path) as xlsx_path:
                sheet_index, data_rows = self.parse_sheet(xlsx_path)
                load_id = f"{datetime.datetime.today().strftime('%Y%m%d_%H%M')}_{get_ascending_letters_within_minute()}_{uuid.uuid4()}"  # noqa: E501
                zip_name = os.path.basename(zip_path)
                unzipped_name = self.inner_file_name

                self.insert_kvp_rows(
                    spark,
                    cat,
                    schema,
                    load_id,
                    zip_name,
                    unzipped_name,
                    self.get_sheet_name(),
                    sheet_index,
                    data_rows,
                )

        logger.info(
            f"load_id:{load_id} loaded {len(data_rows)} enrollment rows from sheet '{self.get_sheet_name()}'"
        )
        return {"data_rows": len(data_rows)}


class TotOrigMeMaOhpEnroll(AbstractLoader):

    def __init__(self):
        super().__init__("MDCR ENROLL AB 1-8_CPS_02ENR_2023.xlsx")

    def get_sheet_name(self):
        return "MDCR ENROLL AB 1_CPS_02ENR"

    def get_first_header_cell_val(self):
        return "Year"

    def get_s3_zip_uri(self):
        return "s3://manipulator-bucket/program_stat_me_total_enroll/CMS Program Statistics - Medicare Total Enrollment.zip"  # noqa E501


def main(*args, **kwargs):
    logger.info("loader main begins")
    cat = kwargs.get("cat", None)
    schema = kwargs.get("schema", None)
    if not cat or not schema:
        cat = sys.argv[1]
        schema = sys.argv[2]
    if not cat or not schema:
        raise ValueError(
            f"Expecting both cat and schema but got {args}, {kwargs}, {sys.argv};"
        )
    logger.info(f"will be using cat:{cat}; schema:{schema};")
    spark = get_spark()
    TotOrigMeMaOhpEnroll().load(spark, cat, schema)
    sql_result = spark.sql("select 1")
    results = [x.asDict() for x in sql_result.toLocalIterator()]
    logger.info(f"loader main end {results}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="loader params")
    parser.add_argument(
        "--cat",
        help="catalog name to use",
        default="b_260723_01_dbr_dbc_cat",
    )
    parser.add_argument(
        "--schema",
        help="schema name to use",
        default="testing_testing",
    )
    args = parser.parse_args()
    main(cat=args.cat, schema=args.schema)
