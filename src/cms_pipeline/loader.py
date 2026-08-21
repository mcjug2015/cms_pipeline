import argparse
import datetime
import os
import re
import sys
import tempfile
import uuid
from typing import Any, Dict, List

from openpyxl import load_workbook
from pyspark.sql import Row, SparkSession
from pyspark.sql.functions import current_timestamp

from src import custom_logging
from src.cms_pipeline.unwrapper import Unwrapper
from src.crutch_migrations.run_crutch_migrations import (
    get_ascending_letters_within_minute,
)
from src.spark_utils import get_spark
from src.utils import convert_to_key, download_s3_zip

logger = custom_logging.setup_logging().getLogger(__name__)


def get_non_empty_cells(row):
    return [str(c).strip() for c in row if c is not None and str(c).strip() != ""]


def is_only_text_cell(non_empty_cells) -> bool:
    if len(non_empty_cells) == 1 and bool(re.search(r"[A-Za-z]", non_empty_cells[0])):
        return True
    return False


def get_sheet_info_dict(toc_worksheet) -> Dict[str, str]:
    result = {}
    for row in toc_worksheet.iter_rows(values_only=True):
        cells_w_values = get_non_empty_cells(row)
        if len(cells_w_values) > 1 and cells_w_values[0] != "Table Name":
            result[cells_w_values[0]] = cells_w_values[1]
    return result


def get_workbook_sheet_info_dict(workbook):
    if "Table of Contents" not in workbook:
        sheet_info_dict = {sheet_name: "" for sheet_name in workbook.sheet_names()}
        logger.info(
            "No 'Table of Contents' sheet in workbook, sheet info will be blank"
        )
    else:
        toc_worksheet = workbook["Table of Contents"]
        sheet_info_dict = get_sheet_info_dict(toc_worksheet)
    return sheet_info_dict


def parse_sheet(
    worksheet,
) -> List[Dict[str, Any]]:
    data_rows: List[Dict[str, Any]] = []
    col_index_to_header_col_name = {}
    for header_row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
        cells = get_non_empty_cells(row)
        if cells and len(cells) > 1:
            for idx, header_cell in enumerate(cells):
                col_index_to_header_col_name[idx] = header_cell
            break

    if not col_index_to_header_col_name:
        return data_rows

    for idx, row in enumerate(
        worksheet.iter_rows(min_row=header_row_idx + 2, values_only=True)
    ):
        cells = get_non_empty_cells(row)
        if (
            len(cells) > 0
            and len(cells[0]) > 0
            and cells[0] != "BLANK"
            and not is_only_text_cell(cells)
        ):
            record = {}
            for idx, value_cell in enumerate(cells):
                record[str(col_index_to_header_col_name[idx])] = str(value_cell)
            data_rows.append(record)

    return data_rows


def insert_kvp_rows(
    spark: SparkSession,
    cat: str,
    schema: str,
    load_id: str,
    zip_name: str,
    unzipped_name: str,
    sheet_name: str,
    sheet_index: int,
    data_rows: List[Dict[str, str]],
) -> int:
    if not data_rows:
        return 0
    rows = []
    for idx, data_row in enumerate(data_rows):
        for the_key, the_val in data_row.items():
            table_key_simple = convert_to_key(the_key)
            rows.append(
                Row(
                    load_id=load_id,
                    zip_name=zip_name,
                    unzipped_name=unzipped_name,
                    sheet_name=sheet_name,
                    sheet_index=sheet_index,
                    table_key=the_key,
                    table_key_simple=table_key_simple,
                    table_row_index=idx,
                    table_val=the_val,
                )
            )
    df = (
        spark.createDataFrame(rows)
        .withColumn("created_at", current_timestamp())
        .withColumn("updated_at", current_timestamp())
    )
    df.writeTo(f"{cat}.{schema}.open_cms_data_kvp").append()
    return len(rows)


def load_cms_workbook(
    spark: SparkSession, cat: str, schema: str, workbook, zip_name, unzipped_name
):
    sheet_info_dict = get_workbook_sheet_info_dict(workbook)
    load_id = f"{datetime.datetime.today().strftime('%Y%m%d_%H%M')}_{get_ascending_letters_within_minute()}_{uuid.uuid4()}"  # noqa: E501
    for sheet_name, _sheet_desc in sheet_info_dict.items():
        data_rows = parse_sheet(workbook[sheet_name])
        insert_kvp_rows(
            spark,
            cat,
            schema,
            load_id,
            zip_name,
            unzipped_name,
            sheet_name,
            list(sheet_info_dict.keys()).index(sheet_name),
            data_rows,
        )


def load_zip_workbook(
    spark: SparkSession, cat: str, schema: str, s3_zip_uri: str
) -> Dict[str, int]:
    with tempfile.TemporaryDirectory(prefix="cms_dl_") as tmp_dir:
        zip_path = download_s3_zip(spark, s3_zip_uri, tmp_dir)
        with Unwrapper().unwrap(zip_path) as xlsx_path:
            zip_name = os.path.basename(zip_path)
            unzipped_name = os.path.basename(xlsx_path)
            return load_cms_workbook(
                spark,
                cat,
                schema,
                load_workbook(xlsx_path, data_only=True, read_only=True),
                zip_name,
                unzipped_name,
            )


def main(*args, **kwargs):  # pragma: no cover
    logger.info("loader main begins")
    cat = kwargs.get("cat", None)
    schema = kwargs.get("schema", None)
    if not cat or not schema:
        cat = sys.argv[1]
        schema = sys.argv[2]
    if not cat or not schema:
        raise ValueError(
            f"Expecting both cat and schema but got {args}, {kwargs}, {sys.argv};"
        )
    logger.info(f"will be using cat:{cat}; schema:{schema};")
    spark = get_spark()
    load_zip_workbook(
        spark,
        cat,
        schema,
        "s3://manipulator-bucket/program_stat_me_total_enroll/CMS Program Statistics - Medicare Total Enrollment ALL.zip",  # noqa: E501
    )
    sql_result = spark.sql("select 1")
    results = [x.asDict() for x in sql_result.toLocalIterator()]
    logger.info(f"loader main end {results}")


if __name__ == "__main__":  # pragma: no cover
    parser = argparse.ArgumentParser(description="loader params")
    parser.add_argument(
        "--cat",
        help="catalog name to use",
        default="b_260723_01_dbr_dbc_cat",
    )
    parser.add_argument(
        "--schema",
        help="schema name to use",
        default="testing_testing",
    )
    args = parser.parse_args()
    main(cat=args.cat, schema=args.schema)
