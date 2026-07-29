import sys
import zipfile
from pathlib import Path
from unittest import mock

import pytest
from openpyxl import Workbook

from src.harness.manipulator import loader


def _build_source_zip(tmp_path):
    workbook = Workbook()
    cover = workbook.worksheets[0]
    cover.title = "Cover"
    cover["A1"] = "cover page, not the target sheet"

    worksheet = workbook.create_sheet("US Total")
    worksheet.append(["CMS Program Statistics"])
    worksheet.append(["SOURCE:", "CMS Chronic Conditions Data Warehouse"])
    worksheet.append(
        [
            "Year",
            "Total Enrollment",
            "% Incr",
            "Orig ME Enroll",
            "% Incr",
            "% of Total",
            "MA/OHP Enroll",
            "% Incr",
            "% of Total",
        ]
    )
    worksheet.append(
        [
            2022,
            "65,000,000",
            "2.1%",
            "35,000,000",
            "1.0%",
            "53.8%",
            "30,000,000",
            "3.5%",
            "46.2%",
        ]
    )
    worksheet.append([2023, 68499235, 3.4, 34000000, -1.0, 49.6, 34499235, 8.0, 50.4])
    worksheet.append(["NOTES:", "Totals may not sum due to rounding."])

    xlsx_path = tmp_path / "MDCR ENROLL AB 1-8_CPS_02ENR_2023.xlsx"
    workbook.save(xlsx_path)

    zip_path = tmp_path / "CMS Program Statistics - Medicare Total Enrollment.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.write(xlsx_path, arcname="MDCR ENROLL AB 1-8_CPS_02ENR_2023.xlsx")
    return str(zip_path)


def test_download_s3_zip_writes_source_content_to_dest_dir(test_spark, tmp_path):
    source_dir = tmp_path / "source"
    source_dir.mkdir()
    source_path = source_dir / "some name.zip"
    source_path.write_bytes(b"dummy zip bytes")

    dest_dir = tmp_path / "dest"
    dest_dir.mkdir()

    dest_path = loader.download_s3_zip(
        test_spark, f"file://{source_path}", str(dest_dir)
    )

    assert dest_path == str(dest_dir / "some name.zip")
    assert Path(dest_path).read_bytes() == b"dummy zip bytes"


def test_parse_sheet_splits_preamble_from_data_rows(tmp_path):
    zip_path = _build_source_zip(tmp_path)
    with loader.TotOrigMeMaOhpEnrollUnwrapper(zip_path).unwrap() as xlsx_path:
        sheet_name, kvp_rows, data_rows = loader.parse_sheet(xlsx_path)

    assert sheet_name == "US Total"
    assert {"table_key": "CMS Program Statistics", "table_value": ""} in kvp_rows
    assert {
        "table_key": "SOURCE:",
        "table_value": "CMS Chronic Conditions Data Warehouse",
    } in kvp_rows
    assert len(data_rows) == 2
    assert data_rows[0] == {
        "row_yr": 2022,
        "tot_enroll": 65000000.0,
        "tot_enroll_pct_increase_prior_yr": 2.1,
        "tot_orig_me_enroll": 35000000.0,
        "tot_orig_me_enroll_pct_increase_prior_yr": 1.0,
        "tot_orig_me_pct_of_tot_enroll": 53.8,
        "tot_ma_ohp_enroll": 30000000.0,
        "tot_ma_ohp_enroll_pct_increase_prior_yr": 3.5,
        "tot_ma_ohp_enroll_pct_of_tot_enroll": 46.2,
    }
    assert data_rows[1]["row_yr"] == 2023
    assert data_rows[1]["tot_enroll"] == 68499235.0


def test_to_float_returns_none_for_none_and_blank_text():
    assert loader._to_float(None) is None
    assert loader._to_float("   ") is None


def test_parse_sheet_skips_blank_rows(tmp_path):
    workbook = Workbook()
    workbook.worksheets[0].title = "Cover"
    worksheet = workbook.create_sheet("US Total")
    worksheet.append(["CMS Program Statistics"])
    worksheet.append([])
    worksheet.append([2023, 68499235, 3.4, 34000000, -1.0, 49.6, 34499235, 8.0, 50.4])
    xlsx_path = tmp_path / "blank_row.xlsx"
    workbook.save(xlsx_path)

    sheet_name, _, data_rows = loader.parse_sheet(str(xlsx_path))

    assert sheet_name == "US Total"
    assert len(data_rows) == 1
    assert data_rows[0]["row_yr"] == 2023


def test_sql_literal_handles_none_and_booleans():
    assert loader._sql_literal(None) == "NULL"
    assert loader._sql_literal(True) == "true"
    assert loader._sql_literal(False) == "false"


def test_insert_kvp_rows_noop_when_empty():
    spark = mock.Mock()
    loader.insert_kvp_rows(
        spark, "cat", "schema", "load_id", "zip", "[]", "unzipped", "sheet", []
    )
    spark.sql.assert_not_called()


def test_insert_data_rows_noop_when_empty():
    spark = mock.Mock()
    loader.insert_data_rows(
        spark, "cat", "schema", "load_id", "zip", "[]", "unzipped", "sheet", []
    )
    spark.sql.assert_not_called()


@mock.patch("src.harness.manipulator.loader.load")
@mock.patch("src.harness.manipulator.loader.get_spark")
def test_main_raises_when_cat_and_schema_missing(
    mock_get_spark, mock_load, monkeypatch
):
    monkeypatch.setattr(sys, "argv", ["loader.py", "", ""])
    with pytest.raises(ValueError):
        loader.main()
    mock_load.assert_not_called()
    mock_get_spark.assert_not_called()
