import os
import shutil
from unittest import mock

from openpyxl.reader.excel import load_workbook

from src import custom_logging
from src.cms_pipeline.loader import load_cms_workbook, load_zip_workbook

logger = custom_logging.setup_logging().getLogger(__name__)
RES_DIR = os.path.join(os.path.dirname(__file__), "res")


def test_load_cms_workbook(migrated_spark, request):
    spark = migrated_spark[0]
    schema = migrated_spark[1]
    logger.info(f"TEST: {request.node.name}; will be using schema {schema};")
    cms_workbook = load_workbook(
        os.path.join(RES_DIR, "MDCR ENROLL AB 15-20_CPS_02ENR_2023.xlsx")
    )
    load_cms_workbook(
        spark, "spark_catalog", schema, cms_workbook, "testing.zip", "testing.xlsx"
    )
    sql_result = spark.sql(
        f"select * from spark_catalog.{schema}.open_cms_data_kvp"
        " where unzipped_name = 'testing.xlsx';"
    )
    results = [x.asDict() for x in sql_result.toLocalIterator()]
    assert len(results) > 0


@mock.patch("src.cms_pipeline.loader.load_cms_workbook")
@mock.patch("src.cms_pipeline.loader.load_workbook")
@mock.patch("src.cms_pipeline.loader.download_s3_zip")
def test_load_zip_workbook(download_s3_zip, load_workbook, load_cms_workbook):
    """load_cms_workbook itself is covered by test_load_cms_workbook above, so it's
    mocked here. This test just validates zip download + nested-zip unwrapping, and
    that load_cms_workbook gets invoked with the unwrapped workbook/zip/file names."""
    inner_file_name = "MDCR ENROLL AB 1-8_CPS_02ENR_2023.xlsx"

    def fake_download_s3_zip(_spark, _s3_uri, dest_dir):
        src_zip = os.path.join(RES_DIR, "nested_total_enroll.zip")
        dest_path = os.path.join(dest_dir, os.path.basename(src_zip))
        shutil.copy(src_zip, dest_path)
        return dest_path

    download_s3_zip.side_effect = fake_download_s3_zip
    load_workbook.return_value = "fake workbook"

    load_zip_workbook(
        None,
        "spark_catalog",
        None,
        "s3://fake-bucket/fake-key.zip",
    )

    download_s3_zip.assert_called_once()
    assert download_s3_zip.call_args.args[0] is None
    assert download_s3_zip.call_args.args[1] == "s3://fake-bucket/fake-key.zip"

    load_workbook.assert_called_once()
    unwrapped_xlsx_path = load_workbook.call_args.args[0]
    assert os.path.basename(unwrapped_xlsx_path) == inner_file_name

    load_cms_workbook.assert_called_once_with(
        None,
        "spark_catalog",
        None,
        "fake workbook",
        "nested_total_enroll.zip",
        inner_file_name,
    )
