import os
from unittest import mock

from openpyxl.reader.excel import load_workbook

from src import custom_logging
from src.cms_pipeline.loader import (
    get_non_empty_cells,
    get_sheet_info_dict,
    get_workbook_sheet_info_dict,
    insert_kvp_rows,
    is_only_text_cell,
    load_zip_workbook,
    parse_sheet,
)

logger = custom_logging.setup_logging().getLogger(__name__)
RES_DIR = os.path.join(os.path.dirname(__file__), "res")


@mock.patch("src.cms_pipeline.loader.convert_to_key", return_value="test converted key")
def test_insert_kvp_rows_success(convert_to_key, migrated_spark, request):
    """
    TODO XXX validate that the schema for these tests is different from the ones test_manipulator and others get
    """
    spark = migrated_spark[0]
    schema = migrated_spark[1]
    logger.info(f"TEST: {request.node.name}; will be using schema {schema};")
    insert_kvp_rows(
        spark,
        cat="spark_catalog",
        schema=schema,
        load_id="test_load_id",
        zip_name="test zip name",
        unzipped_name="test unzipped",
        sheet_name="test sheet name",
        sheet_index=0,
        data_rows=[{"test_key": "test_val"}],
    )
    sql_result = spark.sql(f"select * from spark_catalog.{schema}.open_cms_data_kvp;")
    results = [x.asDict() for x in sql_result.toLocalIterator()]
    assert len(results) == 1
    assert results[0]["load_id"] == "test_load_id"
    assert results[0]["table_key"] == "test_key"
    assert results[0]["table_key_simple"] == "test converted key"
    assert results[0]["table_row_index"] == 0
    assert results[0]["table_val"] == "test_val"
    convert_to_key.assert_called_once()


def test_insert_kvp_rows_no_rows():
    result = insert_kvp_rows(
        None,
        cat=None,
        schema=None,
        load_id=None,
        zip_name=None,
        unzipped_name=None,
        sheet_name=None,
        sheet_index=0,
        data_rows=[],
    )
    assert result == 0


@mock.patch("src.cms_pipeline.loader.download_s3_zip", return_value="/i/am/not/real")
@mock.patch(
    "src.cms_pipeline.loader.Unwrapper.unwrap",
)
@mock.patch("src.cms_pipeline.loader.load_cms_workbook", return_value=1)
@mock.patch("src.cms_pipeline.loader.load_workbook")
def test_load_success(load_workbook, load_cms_workbook, unwrap, download_s3_zip):
    parse_sheet.return_value = (3, [{"test_key": "test_val"}])
    spark = mock.MagicMock(name="spark")
    unwrap.return_value.__enter__.return_value = "/fake/xlsx"

    result = load_zip_workbook(spark, "test_cat", "test_schema", "test_fake_s3_uri")

    assert result == 1
    load_workbook.assert_called_once()
    load_cms_workbook.assert_called_once()
    unwrap.assert_called_once()
    download_s3_zip.assert_called_once()


def test_get_non_empty_cells():
    result = get_non_empty_cells((None, "", "   ", 0, "Year", "  Year  "))

    assert result == ["0", "Year", "Year"]


def test_is_only_text_cell_multiple_cells_returns_false():
    result = is_only_text_cell(["FooterNote", "Metric"])

    assert result is False


def test_is_only_text_cell_single_text_cell_returns_true():
    result = is_only_text_cell(["FooterNote"])

    assert result is True


def test_is_only_text_cell_single_non_text_cell_returns_false():
    result = is_only_text_cell(["12345"])

    assert result is False


def test_get_sheet_info_dict_returns_name_to_description_mapping():
    workbook = load_workbook(os.path.join(RES_DIR, "get_sheet_info_dict_sample.xlsx"))
    result = get_sheet_info_dict(workbook["Table of Contents"])

    # title row (one populated cell), the "Table Name" header row, and the row
    # with only a description (no table name) must all be skipped; whitespace
    # is stripped and extra trailing columns are ignored.
    assert result == {
        "SHEET_A": "Description A",
        "SHEET_B": "Description B",
        "SHEET_C": "Description C",
    }


def test_get_workbook_sheet_info_dict_no_toc():
    workbook = load_workbook(os.path.join(RES_DIR, "parse_sheet_sample.xlsx"))
    result = get_workbook_sheet_info_dict(workbook)

    assert result == {
        "TestSheet": "",
    }


@mock.patch(
    "src.cms_pipeline.loader.get_sheet_info_dict", return_value="testing testing"
)
def test_get_workbook_sheet_info_dict_toc(get_sheet_info_dict):
    workbook = load_workbook(os.path.join(RES_DIR, "get_sheet_info_dict_sample.xlsx"))
    result = get_workbook_sheet_info_dict(workbook)

    assert result == "testing testing"
    get_sheet_info_dict.assert_called_once()


def test_parse_sheet_returns_data_rows():
    workbook = load_workbook(os.path.join(RES_DIR, "parse_sheet_sample.xlsx"))
    data_rows = parse_sheet(workbook["TestSheet"])
    assert data_rows == [
        {"Year": "2023", "Metric": "TotalEnroll", "Value": "100"},
        {"Year": "2024", "Metric": "TotalEnroll", "Value": "200"},
    ]


def test_parse_sheet_returns_no_rows():
    workbook = load_workbook(os.path.join(RES_DIR, "parse_sheet_no_rows_sample.xlsx"))
    data_rows = parse_sheet(workbook["TestSheet"])

    assert data_rows == []
